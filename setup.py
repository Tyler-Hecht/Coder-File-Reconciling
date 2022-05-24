import os
import glob
import shutil
import subprocess
from openpyxl import *
del open
from clearer import clear
from pandas import *

path = os.getcwd()
if os.name == "nt":
    sep = "\\"
else:
    sep = "/"

clear()

input_path = path + "/INPUT/"
output_path = path + "/OUTPUT/"
combined_file = glob.glob(os.path.join(input_path, "*.xlsx"))
third_code = glob.glob(os.path.join(input_path, "*.csv"))

# get input files
if len(combined_file) == 0:
    print("ERROR: no combined file found")
    exit(1)
if len(third_code) == 0:
    print("ERROR: no third coder file found")
    exit(1)
if len(combined_file) > 2:
    print("ERROR: multiple combined files found")
    exit(1)
if len(third_code) > 2:
    print("ERROR: multiple third coder files found")
    exit(1)
combined_file = combined_file[0]
combined_file_name = combined_file.split(sep)[-1]
third_code = third_code[0]
third_code_name = third_code.split(sep)[-1]

os.chdir(input_path)
dts_input_path = path + "/DatavyuToSupercoder/Input/"
dts_output_path = path + "/DatavyuToSupercoder/Output/"

# move file to DTS input
shutil.copyfile(third_code_name, dts_input_path + third_code_name)

# run Java file
os.chdir(path + "/DatavyuToSupercoder/")
subprocess.run("java -jar DatavyuToSupercoder.jar")

# move DTS output file to main output
third_code = glob.glob(os.path.join(dts_output_path, "*.xls"))[0]
third_code_name = third_code.split(sep)[-1]
os.chdir(dts_output_path)
shutil.copyfile(third_code_name, output_path + third_code_name)

# move combined file to output
os.chdir(input_path)
shutil.copyfile(combined_file_name, output_path + combined_file_name)

os.chdir(output_path)
combined_file = glob.glob(os.path.join(output_path, "*.xlsx"))[0]
third_code = glob.glob(os.path.join(output_path, "*.xls"))[0]

os.chdir(path)
with open('config.txt') as f:
    lines = f.readlines()
line2 = lines[5]
trials_and_times = line2.split(",")

trials = []
for trial_and_time in trials_and_times:
    trial_and_time= trial_and_time.strip()
    trials.append(trial_and_time.split(" ")[0])

data = read_excel(third_code, header = 1)
if not (len(trials) == list(data["Code"]).count("B")):
    print("ERROR: Number of trials to reconcile doesn't match number of B's in 3rd coder sheet")
    exit(1)
if not (len(trials) == list(data["Code"]).count("S")):
    print("ERROR: Number of trials to reconcile doesn't match number of S's in 3rd coder sheet")
    exit(1)
# finds the locations of B's in 3rd coder sheet
b_locations = []
s_locations = []
index = 0
while len(s_locations) < len(trials):
    if data["Code"][index] == "B":
        b_locations.append(index)
    if data["Code"][index] == "S":
        s_locations.append(index)
    index += 1
    
os.chdir(output_path)
wb = Workbook()
wb.remove(wb.active)
# create sheets for each trial
for trial in trials:
    name = "Trial " + trial
    wb.create_sheet(name)
wb.save('reconciling.xlsx')

# add 3rd coder trials
for i in range(0, len(trials)):
    index = 2
    for j in range(b_locations[i], s_locations[i]+1):
        name = "Trial " + trials[i]
        sheet = wb[name]
        sheet["A1"] = "3rd coder data for " + name
        sheet["A" + str(index)] = data["Code"][j]
        sheet["B" + str(index)] = data["Onset"][j]
        sheet["C" + str(index)] = data["Offset"][j]
        index += 1
    wb.save('reconciling.xlsx')

wb2 = load_workbook(combined_file)
for i in range(0, 2):
    # finds the B's and S's in the combined file
    sheet2 = wb2.worksheets[i]
    coder = sheet2.title
    locations = {}
    trial_num = 1
    for j in range(1, sheet2.max_row+1):
        cell = sheet2["A" + str(j)]
        if cell.value == "B":
            locations[trial_num] = {}
            locations[trial_num]["B"] = j
        if cell.value == "S":
            locations[trial_num]["S"] = j
            trial_num += 1
    # adds the combined file trials to the sheet
    for sheet in wb:
        index = 2
        trial = int(sheet.title.split()[1])
        b_loc = locations[trial]["B"]
        s_loc = locations[trial]["S"]
        if i == 0:
            sheet["E1"] = coder
        else:
            sheet["I1"] = coder
        for j in range(b_loc, s_loc+1):
            if i == 0:
                sheet["E"+str(index)].value = sheet2["A"+str(j)].value
                sheet["F"+str(index)].value = sheet2["B"+str(j)].value
                sheet["G"+str(index)].value = sheet2["C"+str(j)].value
            else:
                sheet["I"+str(index)].value = sheet2["A"+str(j)].value
                sheet["J"+str(index)].value = sheet2["B"+str(j)].value
                sheet["K"+str(index)].value = sheet2["C"+str(j)].value
            index += 1
        wb.save('reconciling.xlsx')

print("Complete")
