import os
import glob
import shutil
import subprocess
from openpyxl import *
from copy import copy
del open
path = os.getcwd()
combining_path = path + "/combining/"
os.chdir(combining_path)
from combining import combining
os.chdir(path)

tmp = os.getcwd()
if os.name == "nt":
    sep = "\\"
    pre = "python3"
else:
    sep = "/"
    pre = "py"
input_path = path + "/INPUT/"
output_path = path + "/OUTPUT/"
combined_file = glob.glob(os.path.join(input_path, "*.xlsx"))[0].split(sep)[-1]
os.chdir(output_path)

wb = load_workbook("reconciling.xlsx")
wb2 = load_workbook(combined_file)
for i in range(0, 2):
    if i == 0:
        col = "E"
        cols = ["E", "F", "G"]
    else:
        col = "I"
        cols = ["I", "J", "K"]
    # finds the B's and S's in the combined file
    sheet2 = wb2.worksheets[i]
    coder = sheet2.title
    locations = {}
    trial_num = 1
    for j in range(1, sheet2.max_row+1):
        cell = sheet2["A" + str(j)]
        if cell.value == "B":
            locations[trial_num] = {}
            locations[trial_num]["B"] = j
        if cell.value == "S":
            try:
                locations[trial_num]["S"] = j
            except:
                print("ERROR: There's probably missing B in " + coder)
                exit(1)
            trial_num += 1
    # handles the reconciling file
    for sheet in wb:
        trial = int(sheet.title.split()[1])
        coder = sheet[col+"1"].value
        b_loc = locations[trial]["B"]
        s_loc = locations[trial]["S"]
        old_length = s_loc - b_loc + 1
        index = 1
        while sheet[col+str(index)].value:
            index += 1
        new_length = index-2
        diff = new_length - old_length
        if diff < 0:
            wb2[coder].delete_rows(b_loc, -diff)
        if diff > 0:
            wb2[coder].insert_rows(b_loc, diff)
        for key in locations:
            if key > trial:
                locations[key]["B"] += diff
                locations[key]["S"] += diff
        index = 2
        for j in range(b_loc, b_loc + new_length):
            for k in [["A", 0], ["B", 1], ["C", 2]]:
                wb2[coder][k[0]+str(j)].value = sheet[cols[k[1]]+str(index)].value
                wb2[coder][k[0]+str(j)].fill = copy(sheet[cols[k[1]]+str(index)].fill)
                wb2[coder][k[0]+str(j)].font = copy(sheet[cols[k[1]]+str(index)].font)
            index += 1
    wb2.save(combined_file)
        
print("Data added successfully")

# run the combining
shutil.copyfile(combined_file, combining_path + "/input/" + combined_file)
os.chdir(combining_path)
exec(open("catcher.py").read())
os.chdir(combining_path)
combining.main()
os.chdir(combining_path + "/input/")
shutil.copyfile(combined_file, output_path + combined_file)

# check for bad trials
path = tmp
os.chdir(path)
exec(open("recode_finder.py").read())