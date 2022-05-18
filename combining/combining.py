import os
import sys
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"

path = os.getcwd()
config_path = "/".join(os.getcwd().split(sep)[:-1])+"/"
os.chdir(config_path)
with open('config.txt') as f:
    lines = f.readlines()
try:
    study = lines[1][:-1].strip().lower()
except:
    print("No study given in config file. Make sure to put the study type on the second line")
    exit(1)
if not study in ["facetalk", "wls", "awl"]:
    print("Invalid study \"" + study + "\" in config file.")
    exit(1)
os.chdir(path)
'''
Checks if the program is running as an executable or if the program is running through an IDE
Will return the application path of the source file
'''
def get_application_path():
    application_path = ' '
    if getattr(sys, 'frozen', False):#if the program is running as an executable
        application_path = os.path.dirname(sys.executable)
    else: #if program is running through ide
        application_path = os.path.dirname(os.path.abspath(__file__))
    return application_path
'''
Calculate the difference between column B and C in the coder files
'''
def calculate_difference(application_path):
    files = os.listdir(application_path + "/Input/") #get all files in Input folder
    for file in files:
        if(file == ".DS_Store"): #skip .DS_Store file
            continue
        wb = openpyxl.load_workbook(application_path + "/input/" + file) #load excel file
        for sheet in wb: #iterate through each sheet in excel file
            if sheet.title == 'AVERAGES ACROSS CODERS':
                continue
            row_count = 1
            for row in sheet: #for each row in sheet
                difference = 0
                if row[1].value == None and row[2].value == None: #if both row entry are empty
                    difference 
                elif row[1].value == None: #if the first row entry is empty
                    difference = row[2].value - 0
                elif row[2].value == None: #if the second row entry is empty
                    difference = 0 - row[1].value
                else:
                    difference = row[2].value - row[1].value
                sheet.cell(row=row_count, column=4).value = difference #add the difference to the fourth column
                row_count += 1 #continue to the next row
        wb.save(application_path + "/Input/" + file) #save the excel file
'''
 Insert the column header for the coder files
'''       
def insert_col_header(application_path):
    files = os.listdir(application_path + "/Input/")
    for file in files:
        if(file == ".DS_Store"):
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file)
        if study in ["facetalk", "wls"]:
            col_list = ['Left Longest', 'Right Longest', 'Left', 'Right', 'Center', 'Total Look', 'Trial Length', 'Attention']
        elif study in ["awl"]:
            col_list = ['Left Top Longest', 'Right Top Longest', 'Left Bottom Longest', 'Right Bottom Longest', 'Left Top',
                    'Right Top', 'Left Bottom', 'Right Bottom', 'Total Look', 'Trial Length', 'Attention']
        col_count = 5
        list_count = 0
        for sheet in wb:
            if sheet.title == 'AVERAGES ACROSS CODERS':
                continue
            if study in ["facetalk", "wls"]:
                num = 12
            elif study in ["awl"]:
                num = 15
            while col_count <= num:#iterate through all the column title in col_list
                sheet.cell(row=2, column=col_count).value = col_list[list_count]#enter value
                col_count += 1
                list_count += 1
            col_count = 5
            list_count = 0
        wb.save(application_path + "/Input/" + file)
'''
Compute longest looks, sum of each looks, total look, total trial, and attention
Will skip the first 30 seconds of each trial
'''
if study == "facetalk":
    def compute(application_path):
        files = os.listdir(application_path + "/Input/") #get all files of input folder
        for file in files: #iterate through all file
            if(file == ".DS_Store"): #skip .DS_Store file
                continue
            wb = openpyxl.load_workbook(application_path + "/Input/" + file) #load the excel file
            coder_num = 0 #coder 1 or coder 2
            for sheet in wb.worksheets: #for each sheet in excel file
                coder_num += 1
                if sheet.title == 'AVERAGES ACROSS CODERS': #skip average across coder file
                    continue
                row_count = 3 #start row
                l_max = 0 #left longest
                r_max = 0 #right longest
                l_sum = 0 #right 
                r_sum = 0 #left
                c_sum = 0 #center
                start_time = None #start time
                end_time = None #end time
                excel_row_count = 1

                #EDITED FOR FACETALK: immediately calculates time at start of trial (with no buffer of 42 frames)
                for row in sheet.iter_rows():
                    start_look = row[1].value #start value for look
                    end_look = row[2].value #end value for look
                    if row[0].value == None: #if no more values then break from loop
                        break
                    code = row[0].value.strip().upper() #uppercase all code
                    if code == 'B': #if start of trial
                        #count += 1
                        start_time = row[1].value #get start time
                        #print(rt_max, count)
                    elif code == 'L': #if left look
                        if row[3].value > l_max: #check if it is the longest look
                            l_max = row[3].value
                        l_sum = l_sum + row[3].value#add to sum
                    elif code == 'R': #if right look
                        if row[3].value > r_max: #check if it is the longest look
                            r_max = row[3].value
                        r_sum = r_sum + row[3].value #add to sum
                    elif code == 'C': #if center look
                        c_sum = c_sum + row[3].value #add to sum
                    elif code == 'S': #if end of trial
                        end_time = row[1].value #get end time
                        total_look = l_sum + r_sum + c_sum #calculate total_look
                        trial_length = end_time - start_time #calculate trial length
                        attention = total_look / trial_length #calculate attention
                        col_count = 5
                        list_count = 0
                        data = [l_max, r_max, l_sum, r_sum, c_sum, total_look, trial_length, attention]
                        if study in ["facetalk", "wls"]:
                            num = 12
                        elif study in ["awl"]:
                            num = 15
                        while col_count <= num: #add calculated data
                            sheet.cell(row=row_count, column=col_count).value = data[list_count]
                            col_count += 1
                            list_count += 1
                        row_count += 1 #go to next row
                        l_max = 0 #initalize all values to 0
                        r_max = 0
                        l_sum = 0
                        r_sum = 0
                        c_sum = 0
                        start_time = None
                        end_time = None
                    excel_row_count += 1
            wb.save(application_path + "/Input/" + file)
elif study == "wls":
    def compute(application_path):
        files = os.listdir(application_path + "/Input/") #get all files of input folder
        for file in files: #iterate through all file
            if(file == ".DS_Store"): #skip .DS_Store file
                continue
            wb = openpyxl.load_workbook(application_path + "/Input/" + file) #load the excel file
            coder_num = 0 #coder 1 or coder 2
            for sheet in wb.worksheets: #for each sheet in excel file
                coder_num += 1
                if sheet.title == 'AVERAGES ACROSS CODERS': #skip average across coder file
                    continue
                row_count = 3 #start row
                l_max = 0 #left longest
                r_max = 0 #right longest
                l_sum = 0 #right 
                r_sum = 0 #left
                c_sum = 0 #center
                start_time = None #start time
                end_time = None #end time
                excel_row_count = 1
                for row in sheet.iter_rows():
                    if row[0].value == None: #if no more values then break from loop
                        break
                    code = row[0].value.strip().upper() #uppercase all code
                    if code == 'B': #if start of trial
                        #count += 1
                        start_time = row[1].value #get start time
                        #print(rt_max, count)
                    elif code == 'L': #if left look
                        if row[1].value - start_time < 42: #if before 42 frames
                            if (row[1].value - start_time) + row[3].value > 42: #check trial is before 42 frames and goes over 42 frames
                                difference = (row[1].value - start_time) - 42 #caclulate the time that is before 42 frames
                                look = row[3].value + difference #add the difference to get the time that is over 42 frames
                                if look > l_max: #check if it is the longest look
                                    l_max = look
                                l_sum += look #add to sum
                        else:
                            if row[3].value > l_max: #check if it is the longest look
                                l_max = row[3].value
                            l_sum = l_sum + row[3].value#add to sum
                    elif code == 'R': #if right look
                        if row[1].value - start_time < 42: #if before 42 frames
                            if (row[1].value - start_time) + row[3].value > 42: #check trial is before 42 frames and goes over 42 frames
                                difference = (row[1].value - start_time) - 42 #calculate the time that is before 42 frames
                                look = row[3].value + difference #add the difference to get the time that is over 42 frames
                                if look > r_max: #check if it is the longest look
                                    r_max = look
                                r_sum += look #add to sum
                        else:
                            if row[3].value > r_max: #check if it is the longest look
                                r_max = row[3].value
                            r_sum = r_sum + row[3].value #add to sum
                    elif code == 'C': #if center look
                        if row[1].value - start_time < 42: #if before 42 frames
                            if (row[1].value - start_time) + row[3].value > 42: #check trial is before 42 frames and goes over 42 frames
                                difference = (row[1].value - start_time) - 42 #calculate the time that is before 42 frames
                                look = row[3].value + difference #add the difference to get the time that is over 42 frames
                                c_sum += look #add to sum
                        else:
                            c_sum = c_sum + row[3].value #add to sum

                    elif code == 'S': #if end of trial
                        end_time = row[1].value #get end time
                        total_look = l_sum + r_sum + c_sum #calculate total_look
                        trial_length = end_time - start_time #calculate trial length
                        attention = total_look / trial_length #calculate attention
                        col_count = 5
                        list_count = 0
                        data = [l_max, r_max, l_sum, r_sum, c_sum, total_look, trial_length, attention]
                        while col_count <= 12: #add calculated data
                            sheet.cell(row=row_count, column=col_count).value = data[list_count]
                            col_count += 1
                            list_count += 1
                        row_count += 1 #go to next row
                        l_max = 0 #initalize all values to 0
                        r_max = 0
                        l_sum = 0
                        r_sum = 0
                        c_sum = 0
                        start_time = None
                        end_time = None
                    excel_row_count += 1
            wb.save(application_path + "/Input/" + file)
elif study == "awl":
    def compute(application_path):
        files = os.listdir(application_path + "/Input/") #get all files of input folder
        for file in files: #iterate through all file
            if(file == ".DS_Store"): #skip .DS_Store file
                continue
            wb = openpyxl.load_workbook(application_path + "/Input/" + file) #load excel file
            coder_num = 0 #coder 1 or coder 2 
            for sheet in wb.worksheets: #iterate through each sheet 
                coder_num += 1
                if sheet.title == 'AVERAGES ACROSS CODERS': #skup averages across coders sheet
                    continue
                row_count = 3 #start row
                lt_max = 0 #left top longest
                rt_max = 0 #right top longest 
                lb_max = 0 #left bottom longest
                rb_max = 0 #right bottom longest
                lt_sum = 0 #Left Top
                rt_sum = 0 #Right Top
                lb_sum = 0 #Left Bottom
                rb_sum = 0 #Right Bottom
                start_time = None
                end_time = None
                excel_row_count = 1
                for row in sheet.iter_rows():
                    if row[0].value == None: #end loop if end of excel file
                        break
                    code = row[0].value.strip().upper() #make all code uppercase
                    if code == 'B': #if beginning of trial
                        #count += 1
                        start_time = row[1].value #get start time
                        #print(rt_max, count)
                    elif code == 'LT': #if left top look
                        if row[1].value - start_time < 30: #if before 30 seconds 
                            if (row[1].value - start_time) + row[3].value > 30: #check trial is before 30 seconds and goes over 30 seconds
                                difference = (row[1].value - start_time) - 30 #calculate the time that is before 30 seconds
                                look = row[3].value + difference #add the difference to get the time that goes over 30 seconds
                                if look > lt_max: #check if it is the longest look
                                    lt_max = look
                                lt_sum += look #add to sum
                        else:
                            if row[3].value > lt_max: #check if it is the longest look
                                lt_max = row[3].value
                            lt_sum = lt_sum + row[3].value #add to sum
                    elif code == 'RT': #if right top look
                        if row[1].value - start_time < 30: #if before 30 seconds
                            if (row[1].value - start_time) + row[3].value > 30: #check trial is before 30 seconds and goes over 30 seconds
                                difference = (row[1].value - start_time) - 30 #calculate the time that is before 30 seconds
                                look = row[3].value + difference #add the difference to get the time that goes over 30 seconds 
                                if look > rt_max: #check if it is the longest look
                                    rt_max = look
                                rt_sum += look #add to sum
                        else:
                            if row[3].value > rt_max: #check if it is the longest look
                                rt_max = row[3].value
                            rt_sum = rt_sum + row[3].value #add to sum
                    elif code == 'LB': #if left bottom look
                        if row[1].value - start_time < 30: #check trial is before 30 seconds
                            if (row[1].value - start_time) + row[3].value > 30: #check if trial is before 30 seconds and goes over 30 seconds 
                                difference = (row[1].value - start_time) - 30 #calculate the time that is before 30 seconds 
                                look = row[3].value + difference #add the difference to get the time that goes over 30 seconds 
                                if look > lb_max: #check if it is the longest look
                                    lb_max = look
                                lb_sum += look #add to sum
                        else:
                            if row[3].value > lb_max: #check if it is the longest look
                                lb_max = row[3].value
                            lb_sum = lb_sum + row[3].value #add to sum
                    elif code == 'RB': #if right bottom look
                        if row[1].value - start_time < 30: #check trial is before 30 seconds
                            if (row[1].value - start_time) + row[3].value > 30: #check if trial is before 30 seconds and goes over 30 seconds
                                difference = (row[1].value - start_time) - 30 #calculate the time that is before 30 seconds 
                                look = row[3].value + difference #add the difference to get the time that goes over 30 seconds 
                                if look > rb_max: #check if it is the longest look
                                    rb_max = look
                                rb_sum += look #add to sum
                        else:
                            if row[3].value > rb_max: #check if it is the longest loom
                                #print(row[3].value, "!!!!!!!!!!")
                                rb_max = row[3].value 
                            rb_sum = rb_sum + row[3].value #add to sum
                    elif code == 'S': #if it is the end of trial
                        end_time = row[1].value #get end time
                        total_look = lt_sum + rt_sum + lb_sum + rb_sum #calculate total look
                        trial_length = end_time - start_time #calculate trial length
                        attention = total_look / trial_length #calculate attention
                        col_count = 5
                        list_count = 0
                        data = [lt_max, rt_max, lb_max, rb_max, lt_sum, rt_sum, lb_sum, rb_sum, total_look, trial_length, attention]
                        while col_count <= 15: #insert data to excel file
                            sheet.cell(row=row_count, column=col_count).value = data[list_count]
                            col_count += 1
                            list_count += 1
                        row_count += 1 # go to next row 
                        lt_max = 0 #initialize all values back to 0
                        rt_max = 0
                        lb_max = 0
                        rb_max = 0
                        lt_sum = 0
                        rt_sum = 0
                        lb_sum = 0
                        rb_sum = 0
                        start_time = None
                        end_time = None
                    excel_row_count += 1
            wb.save(application_path + "/Input/" + file)
'''
Insert column headers for AVERAGE ACROSS CODERS excel sheet
'''
def insert_dis_col_header(application_path):
    files = os.listdir(application_path + "/Input/")
    for file in files:
        if(file == ".DS_Store"):
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file)
        if study in ["facetalk", "wls"]:
            col_list = ['Left Longest', 'Left Discrep','Right Longest', 'Rt Discrep', 'Left',
                        'L dis', 'Right', 'R dis', 'Center', 'C dis','Total Look', 'Total discr', 'Trial Length',
                        'Length Discr', 'Attention']
        elif study in ["awl"]:
            col_list = ['Left Top Longest', 'Left Top Longest Dis','Right Top Longest', 'Rt Top Longest Dis', 'Left Bot Longest',
                    'Left Bot Longest Dis', 'Right Bottom Longest', 'RT Bot Longest Dis', 'Left Top', 'LT Dis',
                    'Right Top', 'RT Dis', 'Left Bottom', 'LB Dis', 'Right Bottom', 'RB Dis', 'Total Look', 'Total Dis', 'Trial Length',
                    'Trial Dis', 'Attention']
        col_count = 1
        list_count = 0
        if len(wb.worksheets) == 2:#check if there is only two sheets
            wb.create_sheet('AVERAGES ACROSS CODERS')#create new sheet if it doesn't exisit 
        yellowFill = PatternFill(start_color = 'FFFF00',
                                 end_color = 'FFFF00',
                                 fill_type = 'solid')
        for sheet in wb:
            if sheet.title == 'AVERAGES ACROSS CODERS':
                if study in ["facetalk", "wls"]:
                    num = 15
                elif study in ["awl"]:
                    num = 21
                while col_count <= num:#add column headers to excel file
                    sheet.cell(row=1, column=col_count).value = col_list[list_count]
                    cur_cell = sheet.cell(1, col_count)
                    if col_count % 2 == 0:
                        cur_cell.fill = yellowFill
                    col_count += 1
                    list_count += 1
                col_count = 1
                list_count = 0
        wb.save(application_path + "/Input/" + file)
'''
Calculate discrepencies between the two coders
'''    
if study in ["facetalk", "wls"]:   
    def compute_dis(application_path):
        files = os.listdir(application_path + "/Input/")
        for file in files:
            if(file == ".DS_Store"):
                continue
            wb = openpyxl.load_workbook(application_path + "/Input/" + file)
            sheet1 = wb.worksheets[0] #coder 1
            sheet2 = wb.worksheets[1] #coder 2
            sheet3 = wb.worksheets[2] #average across coders
            l_max = 0 #left longest
            l_max_dis = 0 #left longest disp
            r_max = 0 #right longest
            r_max_dis = 0 #right longest disp
            l_sum = 0 #left 
            l_sum_dis = 0 #left disp
            r_sum = 0 #right 
            r_sum_dis = 0 #right disp 
            c_sum = 0 #center
            c_sum_dis = 0 #center disp
            total_look_sum = 0 #total look
            total_look_dis = 0 #total look disp
            trial_length_sum = 0 #trial length 
            trial_length_dis = 0 #trial length disp
            attention_sum = 0 #attenion
            prev_row_count = 1
            new_row_count = 2
            yellowFill = PatternFill(start_color = 'FFFF00',
                                    end_color = 'FFFF00',
                                    fill_type = 'solid')
            redFill = PatternFill(start_color = 'FF0000',
                                end_color = 'FF0000',
                                fill_type = 'solid')
            for row1, row2 in zip(sheet1, sheet2):
                if prev_row_count == 1 or prev_row_count == 2:
                    prev_row_count += 1
                    continue
                if row1[4].value == None or row2[4].value == None:
                    break
                #calculate left longest and left longest disp
                l_max = (row1[4].value + row2[4].value) / 2
                sheet3.cell(row=new_row_count, column = 1).value = l_max
                l_max_dis = row1[4].value - row2[4].value
                sheet3.cell(row=new_row_count, column = 2).value = l_max_dis
                cur_cell = sheet3.cell(new_row_count, 2)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right longest and right longest disp
                r_max = (row1[5].value + row2[5].value) / 2
                sheet3.cell(row=new_row_count, column = 3).value = r_max
                r_max_dis = row1[5].value - row2[5].value
                sheet3.cell(row=new_row_count, column = 4).value = r_max_dis
                cur_cell = sheet3.cell(new_row_count, 4)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate left and left disp
                l_sum = (row1[6].value + row2[6].value) / 2
                sheet3.cell(row=new_row_count, column = 5).value = l_sum
                l_sum_dis = row1[6].value - row2[6].value
                sheet3.cell(row=new_row_count, column = 6).value = l_sum_dis
                cur_cell = sheet3.cell(new_row_count, 6)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right and right disp
                r_sum = (row1[7].value + row2[7].value) / 2
                sheet3.cell(row=new_row_count, column = 7).value = r_sum
                r_sum_dis = row1[7].value - row2[7].value
                sheet3.cell(row=new_row_count, column = 8).value = r_sum_dis
                cur_cell = sheet3.cell(new_row_count, 8)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate center and center disp
                c_sum = (row1[8].value + row2[8].value) / 2
                sheet3.cell(row=new_row_count, column = 9).value = c_sum
                c_sum_dis = row1[8].value - row2[8].value
                sheet3.cell(row=new_row_count, column = 10).value = c_sum_dis
                cur_cell = sheet3.cell(new_row_count, 10)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate total look and total look disp
                total_look_sum = (row1[9].value + row2[9].value) / 2
                sheet3.cell(row=new_row_count, column = 11).value = total_look_sum
                total_look_dis = row1[9].value - row2[9].value
                sheet3.cell(row=new_row_count, column = 12).value = total_look_dis
                cur_cell = sheet3.cell(new_row_count, 12)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate trial length and trial length disp
                trial_length_sum = (row1[10].value + row2[10].value) / 2
                sheet3.cell(row=new_row_count, column = 13).value = trial_length_sum
                trial_length_dis = row1[10].value - row2[10].value
                sheet3.cell(row=new_row_count, column = 14).value = trial_length_dis
                cur_cell = sheet3.cell(new_row_count, 14)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                
                attention_sum = (row1[11].value + row2[11].value) / 2
                sheet3.cell(row=new_row_count, column = 15).value = attention_sum
                #initalize values to 0
                l_max = 0 #
                l_max_dis = 0 #
                r_max = 0 #
                r_max_dis = 0 #
                l_sum = 0 #
                l_sum_dis = 0 #
                r_sum = 0 #
                r_sum_dis = 0 #
                c_sum = 0 #
                c_sum_dis = 0 #
                total_look_sum = 0 #
                total_look_dis = 0 #
                trial_length_sum = 0
                trial_length_dis = 0
                attention_sum = 0
                new_row_count += 1
            wb.save(application_path + "/Input/" + file)
elif study in ["awl"]:
    def compute_dis(application_path):
        files = os.listdir(application_path + "/Input/")
        for file in files:
            if(file == ".DS_Store"):
                continue
            wb = openpyxl.load_workbook(application_path + "/Input/" + file)
            sheet1 = wb.worksheets[0] #coder 1
            sheet2 = wb.worksheets[1] #coder 2
            sheet3 = wb.worksheets[2] #average across coders
            lt_max_sum = 0 #Left Top longest
            lt_max_dis = 0 #Left Top longest dis
            rt_max_sum = 0 #right top longest
            rt_max_dis = 0 #right top longest dis
            lb_max_sum = 0 #left bottom longest 
            lb_max_dis = 0 #left bottom longest dis
            rb_max_sum = 0 #right bottom longest
            rb_max_dis = 0 #right bottom longest dis
            lt_sum = 0 #left top
            lt_dis = 0 #left top dis
            rt_sum = 0 #right top
            rt_dis = 0 #right top dis
            lb_sum = 0 #left bottom
            lb_dis = 0 #left bottom dis
            rb_sum = 0 #right bottom
            rb_dis = 0 #right bottom dis
            total_look_sum = 0 #total look
            total_look_dis = 0 #total look dis
            trial_length_sum = 0 #trial length 
            trial_length_dis = 0 #trial length disp
            attention_sum = 0
            prev_row_count = 1
            new_row_count = 2
            yellowFill = PatternFill(start_color = 'FFFF00',
                                    end_color = 'FFFF00',
                                    fill_type = 'solid')
            redFill = PatternFill(start_color = 'FF0000',
                                end_color = 'FF0000',
                                fill_type = 'solid')
            for row1, row2 in zip(sheet1, sheet2):
                if prev_row_count == 1 or prev_row_count == 2:
                    prev_row_count += 1
                    continue
                if row1[4].value == None or row2[4].value == None:
                    break
                #calculate left top longest and left top longest disp
                lt_max_sum = (row1[4].value + row2[4].value) / 2
                sheet3.cell(row=new_row_count, column = 1).value = lt_max_sum
                lt_max_dis = row1[4].value - row2[4].value
                sheet3.cell(row=new_row_count, column = 2).value = lt_max_dis
                cur_cell = sheet3.cell(new_row_count, 2)
                if cur_cell.value > 15 or cur_cell.value < -15: #color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right top longest and right top longest disp
                rt_max_sum = (row1[5].value + row2[5].value) / 2
                sheet3.cell(row=new_row_count, column = 3).value = rt_max_sum
                rt_max_dis = row1[5].value - row2[5].value
                sheet3.cell(row=new_row_count, column = 4).value = rt_max_dis
                cur_cell = sheet3.cell(new_row_count, 4)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate left bottom longest and left bottom longest disp
                lb_max_sum = (row1[6].value + row2[6].value) / 2
                sheet3.cell(row=new_row_count, column = 5).value = lb_max_sum
                lb_max_dis = row1[6].value - row2[6].value
                sheet3.cell(row=new_row_count, column = 6).value = lb_max_dis
                cur_cell = sheet3.cell(new_row_count, 6)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right bottom longest and right bottom longest disp
                rb_max_sum = (row1[7].value + row2[7].value) / 2
                sheet3.cell(row=new_row_count, column = 7).value = rb_max_sum
                rb_max_dis = row1[7].value - row2[7].value
                sheet3.cell(row=new_row_count, column = 8).value = rb_max_dis
                cur_cell = sheet3.cell(new_row_count, 8)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate left top and left top disp
                lt_sum = (row1[8].value + row2[8].value) / 2
                sheet3.cell(row=new_row_count, column = 9).value = lt_sum
                lt_dis = row1[8].value - row2[8].value
                sheet3.cell(row=new_row_count, column = 10).value = lt_dis
                cur_cell = sheet3.cell(new_row_count, 10)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right top and right top disp
                rt_sum = (row1[9].value + row2[9].value) / 2
                sheet3.cell(row=new_row_count, column = 11).value = rt_sum
                rt_dis = row1[9].value - row2[9].value
                sheet3.cell(row=new_row_count, column = 12).value = rt_dis
                cur_cell = sheet3.cell(new_row_count, 12)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate left bottom and left bottom disp
                lb_sum = (row1[10].value + row2[10].value) / 2
                sheet3.cell(row=new_row_count, column = 13).value = lb_sum
                lb_dis = row1[10].value - row2[10].value
                sheet3.cell(row=new_row_count, column = 14).value = lb_dis
                cur_cell = sheet3.cell(new_row_count, 14)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate right bottom and right bottom disp
                rb_sum = (row1[11].value + row2[11].value) / 2
                sheet3.cell(row=new_row_count, column = 15).value = rb_sum
                rb_dis = row1[11].value - row2[11].value
                sheet3.cell(row=new_row_count, column = 16).value = rb_dis
                cur_cell = sheet3.cell(new_row_count, 16)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate total look and total look disp
                total_look_sum = (row1[12].value + row2[12].value) / 2
                sheet3.cell(row=new_row_count, column = 17).value = total_look_sum
                total_look_dis = row1[12].value - row2[12].value
                sheet3.cell(row=new_row_count, column = 18).value = total_look_dis
                cur_cell = sheet3.cell(new_row_count, 18)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                #calculate trial length and trial length disp
                trial_length_sum = (row1[13].value + row2[13].value) / 2
                sheet3.cell(row=new_row_count, column = 19).value = trial_length_sum
                trial_length_dis = row1[13].value - row2[13].value
                sheet3.cell(row=new_row_count, column = 20).value = trial_length_dis
                cur_cell = sheet3.cell(new_row_count, 20)
                if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                    cur_cell.fill = redFill
                else:   
                    cur_cell.fill = yellowFill
                
                attention_sum = (row1[14].value + row2[14].value) / 2
                sheet3.cell(row=new_row_count, column = 21).value = attention_sum
                
                lt_max_sum = 0 #
                lt_max_dis = 0 #
                rt_max_sum = 0 #
                rt_max_dis = 0 #
                lb_max_sum = 0 #
                lb_max_dis = 0 #
                rb_max_sum = 0 #
                rb_max_dis = 0 #
                lt_sum = 0 #
                lt_dis = 0 #
                rt_sum = 0 #
                rt_dis = 0 #
                lb_sum = 0 #
                lb_dis = 0 #
                rb_sum = 0 #
                rb_dis #
                total_look_sum = 0 #
                total_look_dis = 0 #
                trial_length_sum = 0
                trial_length_dis = 0
                attention_sum = 0
                new_row_count += 1
            wb.save(application_path + "/Input/" + file)

def main():
    application_path = get_application_path()
    calculate_difference(application_path)
    insert_col_header(application_path)
    compute(application_path)
    insert_dis_col_header(application_path)
    compute_dis(application_path)
    print('Averages across coders completed!')

if __name__ == "__main__":
    main()
