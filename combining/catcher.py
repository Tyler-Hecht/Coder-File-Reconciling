import os
import glob
from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
del open

# "catcher in the py" lol
def main():
    if os.name == "nt":
        sep = "\\"
    else:
        sep = "/"
        
    path = os.getcwd() + "/Input/"
    config_path = "/".join(os.getcwd().split(sep)[:-1])+"/"
    os.chdir(config_path)
    with open('config.txt') as f:
        lines = f.readlines()
    try:
        study = lines[1][:-1]
    except:
        print("No study given in config file. Make sure to put the study type on the second line")
        exit(1)
    if not study.strip().lower() in ["facetalk", "wls", "awl"]:
        print("Invalid study \"" + study[:-1] + "\" in config file.")
        exit(1)
    try:
        num_trials = int(lines[3].strip())
    except:
        print("No number of trials in config file. Make sure to put the number of trials on the fourth line")
        exit(1)

    os.chdir

    file = glob.glob(os.path.join(path, "*xlsx"))
    file_name = file[0].split(sep)[-1]

    os.chdir(path)

    wb = load_workbook(file_name)

    # creates the row given the row number, look, onset, and offset
    def recreate_line(i, look, on, off) -> str:
        error_row = str(i+1) + " "
        if i+1 < 10:
            error_row += " "
        if i+1 < 100:
            error_row += " "
        error_row += look + " "
        if len(look) == 1:
            error_row += " "
        error_row += str(on) + " "
        if on < 10000:
            error_row +=  " "
        if on < 1000:
            error_row +=  " "
        error_row += str(off)
        return error_row

    def error_region(i, sheet, row_num) -> str:
        error_region = ""
        # handles edge cases
        if i == 0:
            area = [i, i+1]
        elif i == row_num - 1:
            area = [i-1, i]
        else:
            area = [i-1, i, i+1]
        # gets the relevant row and two surrounding rows
        for j in area:
            # Note: spaces are added so the widths are the same for every line
            row = list(list(sheet)[j])
            error_region += str(j+1) + " "
            # adds spaces if there aren't three digits in the row number
            if j+1 < 100:
                error_region += " "
            if j+1 < 10:
                error_region += " "
            for cell in row[0:3]:
                if not cell.value:
                    # I think this fills in the part of the line if there isn't a look?
                    if cell == row[0]:
                        error_region += "  "
                    else:
                        error_region += "      "
                else:
                    # adds spaces if there aren't five digits in the on/offset
                    if cell == row[1] or cell == row[2]:
                        if cell.value < 10000:
                            error_region += " "
                        if cell.value < 1000:
                            error_region += " "
                    error_region += str(cell.value) + " "
                    if cell == row[0]:
                        # adds a space if the look is only one letter
                        if len(cell.value) == 1:
                            error_region += " "
            error_region += "\n"
        return error_region

    # attempts to fix the errors, returning False if successful
    def fix_missing(look: str, i, sheet, set: str) -> bool:
        def is_close(val1, val2, tolerance) -> bool:
            diff = val1 - val2
            return abs(diff) <= tolerance
        if "1" in sheet.title:
            other_sheet_index = 1
        else:
            other_sheet_index = 0
        other_sheet = list(wb)[other_sheet_index]
        other_row_num = 0
        for row in other_sheet:
            other_row_num += 1
            if row[0].value == look:
                if set == "on" and is_close(row[2].value, sheet[i+1][2].value, 3):
                    print("Suggested fix:\n")
                    print(recreate_line(i, look, row[1].value, sheet[i+1][2].value))
                    print()
                    print("Because other coder has:\n")
                    print(error_region(other_row_num-1, other_sheet, len(list(other_sheet))))
                    approval = input("\nApprove fix? (y/n)")
                    if approval in ["y", "Y", "yes", "Yes", "YES"]:
                        # highlights changed cell
                        sheet[i+1][1].fill = PatternFill(start_color = 'FF0000',
                                    end_color = 'FF0000',
                                    fill_type = 'solid')
                        sheet[i+1][1].value = row[1].value
                        return False
                    else:
                        return True
                elif set == "off" and is_close(row[1].value, sheet[i+1][1].value, 3):
                    print("Suggested fix:\n")
                    print(recreate_line(i, look, sheet[i+1][1].value, row[2].value))
                    print()
                    print("Because other coder has:\n")
                    print(error_region(other_row_num-1, other_sheet, len(list(other_sheet))))
                    approval = input("\nApprove fix? (y/n)")
                    if approval in ["y", "Y", "yes", "Yes", "YES"]:
                        # highlights changed cell
                        sheet[i+1][2].fill = PatternFill(start_color = 'FF0000',
                                    end_color = 'FF0000',
                                    fill_type = 'solid')
                        sheet[i+1][2].value = row[2].value
                        return False
                    else:
                        return True

        return True

    # catches errors, highlights B cells, and indexes trials
    error = False
    for sheet in wb:
        b = 0
        s = 0
        row_num = 0
        b_fill = PatternFill(start_color = 'FFD3AA', end_color = 'FFD3AA', fill_type = 'solid')
        # Delete AAC sheet
        if sheet.title == 'AVERAGES ACROSS CODERS':
            wb.remove(sheet)
            wb.save(path + file_name)
            continue
        print("Analyzing " + sheet.title)
        # quick preliminary scan
        for row in sheet:
            if row[0].value == "B":
                b += 1
                # highlights cells and indexes trials
                row[0].fill = b_fill
                sheet["P" + str(row_num + 1)] = b
                sheet["P" + str(row_num + 1)].fill = b_fill
            if row[0].value == "S":
                s += 1
            row_num += 1
            
        # catches errors
        for i in range(0, row_num):
            row = list(sheet)[i]
            # checks offsets and onsets
            if row[0].value in ["B", "S"]:
                if not row[1].value:
                    print("\n" + sheet.title + " missing onset in row " + str(i+1) + "\n")
                    print(error_region(i, sheet, row_num))
                    possible_error = fix_missing(row[0].value, i, sheet, "on")
                    if possible_error:
                        error = True
                if row[2].value:
                    print("\n" + sheet.title + " has offset in row " + str(i+1) + "\n")
                    print(error_region(i, sheet, row_num))
                    # suggests to delete offset for B or S
                    print("Suggested fix:")
                    print(recreate_line(i, row[0].value, row[1].value, ""))
                    approval = input("\nApprove fix? (y/n)")
                    if approval in ["y", "Y", "yes", "Yes", "YES"]:
                        # highlights changed cell
                        row[2].fill = PatternFill(start_color = 'FF0000',
                                    end_color = 'FF0000',
                                    fill_type = 'solid')
                        row[2].value = ""
                    else:
                        error = True
            elif row[0].value in ["R", "L", "C", "RT", "RB", "LT", "LB"]:
                if not row[1].value:
                    print("\n" + sheet.title + " missing onset in row " + str(i+1) + "\n")
                    print(error_region(i, sheet, row_num))
                    possible_error = fix_missing(row[0].value, i, sheet, "on")
                    if possible_error:
                        error = True
                if not row[2].value:
                    print("\n" + sheet.title + " missing offset in row " + str(i+1) + "\n")
                    print(error_region(i, sheet, row_num))
                    possible_error = fix_missing(row[0].value, i, sheet, "off")
                    if possible_error:
                        error = True
            else:
                print("\n" + sheet.title + " has unrecognized look in row " + str(i+1) + "\n")
                print(error_region(i, sheet, row_num))
                error = True
            # makes sure every S has a B following it (except for the last one)
            if row[0].value == "S" and i != row_num - 1:
                if list(sheet)[i+1][0].value != "B":
                    print("\n" + sheet.title + " has an S that isn't followed by a B in row " + str(i+1) + "\n")
                    print(error_region(i+1, sheet, row_num))
                    error = True
            if row[0].value == "B" and i != 0:
                if list(sheet)[i-1][0].value != "S":
                    print("\n" + sheet.title + " has an B that isn't preceded by a S in row " + str(i+1) + "\n")
                    print(error_region(i-1, sheet, row_num))
                    error = True
        # makes sure the number of trials is correct
        if b != num_trials or s != num_trials:
            print("\n" + sheet.title + " has incorrect number of trials.\n" + str(b) + " B\n" + str(s) + " S\n" + "Should have " + str(num_trials) + " of each." + "\n")       
            error = True
        wb.save(path + file_name)
    if error:
        exit(1)
