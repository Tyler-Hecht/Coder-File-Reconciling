import os
import glob
from openpyxl import load_workbook
from pandas import *
from openpyxl.styles import Color, PatternFill, Font, Border

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"

# gets the study information
with open('config.txt') as f:
    lines = f.readlines()
try:
    study = lines[1][:-1]
except:
    print("No study given in config file. Make sure to put the study type on the second line")
    exit(1)
if not study.strip().lower() in ["facetalk", "wls", "awl"]:
    print("Invalid study \"" + study + "\" in config file.")
    exit(1)
else:
    study = study.strip().lower()
try:
    num_trials = int(lines[3].strip())
except:
    print("No number of trials in config file. Make sure to put the number of trials on the fourth line")
    exit(1)
if study in ["facetalk", "wls"]:
    cols = 15
elif study in ["awl"]:
    cols = 21

path = os.getcwd()
input_path_1 = path + "/OUTPUT/"
input_path_2 = path + "/DatavyuToSupercoder/Output/"

output_file = glob.glob(os.path.join(input_path_1, "*.xlsx"))[0].split(sep)[-1]
trials_file = glob.glob(os.path.join(input_path_2, "*.xls"))[0].split(sep)[-1]

os.chdir(input_path_1)
output_wb = load_workbook(output_file)
output_sheet = list(list(output_wb)[2])

os.chdir(input_path_2)
trials_wb = read_excel(trials_file, skiprows=[0])
trial_times = trials_wb["Start Time (in elapsed time - Datavyu coding)"]

bad_trials = set()
redFill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
redFill2 = PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
# find the trials marked red
for i in range(1, num_trials+1):
    row = output_sheet[i]
    for j in range(0, cols):
        cell = row[j]
        if cell.fill == redFill or cell.fill == redFill2:
            bad_trials.add(i)

if len(bad_trials) > 1:
    print("\n" + str(len(bad_trials)) + " bad trials still found. Check averages across coders in OUTPUT folder.")
elif len(bad_trials) == 1:
    print("1 bad trial still found. Check averages across coders in OUTPUT folder.")
else:
    print("Complete: all trials fixed")
